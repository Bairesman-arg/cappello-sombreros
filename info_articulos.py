# -*- coding: utf-8 -*-
import datetime
import calendar
import pandas as pd
import streamlit as st
from sqlalchemy import text
from models import engine
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import numpy as np
import config

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def generar_excel_articulos(agrupado, total_remitos_mes, total_articulos_mes, total_venta_mes, total_utilidad_mes):
    output = io.BytesIO()
    df_excel = pd.DataFrame({
        "Posición": agrupado["Posición"],
        "Artículo": agrupado["Artículo"],
        "Descripción": agrupado["Descripción"],
        "Rubro": agrupado["Rubro"],
        "Cant. Remitos": agrupado["Cant. Remitos"],
        "Cant. Artículos": agrupado["Cant. Artículos"],
        "Utilidad ($)": agrupado["utilidad_articulo"].round(2),
        "Margen de Ganancia (%)": agrupado["Margen de Ganancia (%)"]
    })
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Ranking por Artículo")
        ws = writer.sheets["Ranking por Artículo"]
        
        # Estilos openpyxl
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo pastel tenue
        bold_font = Font(bold=True)

        # Aplicar estilo al encabezado (Fila 1)
        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = bold_font

        # Formatear datos
        for row in range(2, len(df_excel) + 2):
            ws.cell(row=row, column=1).number_format = '#,##0'
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=8).number_format = '0.00"%"'
        
        # Fila final de Totales en Negrita
        total_margen_mes = (total_utilidad_mes / total_venta_mes * 100.0) if total_venta_mes > 0 else 0.0
        tot_row = len(df_excel) + 3
        c1 = ws.cell(row=tot_row, column=1, value="Total del Mes")
        c5 = ws.cell(row=tot_row, column=5, value=total_remitos_mes)
        c5.number_format = '#,##0'
        c6 = ws.cell(row=tot_row, column=6, value=total_articulos_mes)
        c6.number_format = '#,##0'
        c7 = ws.cell(row=tot_row, column=7, value=total_utilidad_mes)
        c7.number_format = '"$"#,##0.00'
        c8 = ws.cell(row=tot_row, column=8, value=total_margen_mes)
        c8.number_format = '0.00"%"'

        top_border = Border(top=Side(style='thin'))
        for col_idx in range(1, 9):
            cell = ws.cell(row=tot_row, column=col_idx)
            cell.font = bold_font
            cell.border = top_border
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
    return output.getvalue()

def info_articulos_ranking():
    st.header("Informes - Ranking por Artículo")
    st.markdown('<div style="font-size: 0.85rem; color: #00E676; font-weight: 500; margin-top: -0.4rem; margin-bottom: 0.8rem;">El Análisis Gráfico se encuentra al final de la página</div>', unsafe_allow_html=True)
    st.markdown("---")

    # Fecha actual para predeterminar la selección
    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    # Controles de selección de Mes y Año
    col_mes, col_anio, _ = st.columns([2, 2, 4], gap="small")
    
    with col_mes:
        mes_nombre = st.selectbox(
            "Mes",
            options=MESES,
            index=current_month - 1,
            key="info_articulos_mes"
        )
        mes_num = MESES.index(mes_nombre) + 1

    with col_anio:
        min_year = min(2026, current_year)
        max_year = max(2050, current_year)
        anios_disponibles = list(range(min_year, max_year + 1))
        index_anio = anios_disponibles.index(current_year)
        anio_num = st.selectbox(
            "Año",
            options=anios_disponibles,
            index=index_anio,
            key="info_articulos_anio"
        )

    # Rango de fechas del mes completo seleccionado
    _, ult_dia = calendar.monthrange(anio_num, mes_num)
    fecha_inicio = datetime.date(anio_num, mes_num, 1)
    fecha_fin = datetime.date(anio_num, mes_num, ult_dia)

    # Consulta a base de datos de remitos con fecha_retiro en el mes seleccionado
    try:
        with engine.begin() as conn:
            query = text("""
                SELECT 
                    r.id AS remito_id,
                    r.fecha_retiro,
                    a.id AS articulo_id,
                    COALESCE(a.nro_articulo, '') AS nro_articulo,
                    COALESCE(a.descripcion, '') AS descripcion,
                    COALESCE(rub.nombre_rubro, 'Sin Rubro') AS rubro_nombre,
                    COALESCE(r.porc_dto, c.porc_dto, 0) AS porc_dto,
                    ri.entregados,
                    COALESCE(ri.devueltos, 0) AS devueltos,
                    COALESCE(ri.precio_real_item, a.precio_real, 0) AS precio_real,
                    COALESCE(a.costo, 0) AS costo
                FROM remitos r
                JOIN clientes c ON r.cliente_id = c.id
                JOIN remito_items ri ON r.id = ri.remito_id
                JOIN articulos a ON ri.articulo_id = a.id
                LEFT JOIN rubros rub ON a.rubro_id = rub.id
                WHERE r.fecha_retiro IS NOT NULL 
                  AND r.fecha_retiro >= :start_date 
                  AND r.fecha_retiro <= :end_date
            """)
            df_raw = pd.read_sql(query, conn, params={"start_date": fecha_inicio, "end_date": fecha_fin})
    except Exception as e:
        st.error(f"Error al consultar la base de datos: {e}")
        return

    if df_raw.empty:
        st.info(f"ℹ️ No existen Remitos con ventas registradas en {mes_nombre} {anio_num}.")
        st.metric("Total Utilidad del Mes", "$ 0.00")
        return

    # Aplicar cálculo de vendidos, venta total y utilidad ítem por ítem
    def calcular_item(row):
        entregados = int(row["entregados"]) if pd.notna(row["entregados"]) else 0
        devueltos = int(row["devueltos"]) if pd.notna(row["devueltos"]) else 0
        vendidos = max(0, entregados - devueltos)

        p_real = float(row["precio_real"]) if pd.notna(row["precio_real"]) else 0.0
        porc_dto = float(row["porc_dto"]) if pd.notna(row["porc_dto"]) else 0.0
        costo = float(row["costo"]) if pd.notna(row["costo"]) else 0.0

        if p_real > 0 and vendidos > 0:
            p_dto = p_real * (1.0 - (porc_dto / 100.0))
            venta_total = p_dto * vendidos
            utilidad = (p_dto - costo) * vendidos
        else:
            venta_total = 0.0
            utilidad = 0.0

        return pd.Series({"vendidos": vendidos, "venta_total": venta_total, "utilidad": utilidad})

    res_items = df_raw.apply(calcular_item, axis=1)
    df_raw["vendidos"] = res_items["vendidos"]
    df_raw["venta_total"] = res_items["venta_total"]
    df_raw["utilidad"] = res_items["utilidad"]

    # Agrupar por Artículo (nro_articulo, descripcion, rubro_nombre)
    agrupado = df_raw.groupby(["articulo_id", "nro_articulo", "descripcion", "rubro_nombre"]).agg(
        cant_articulos=("vendidos", "sum"),
        venta_articulo=("venta_total", "sum"),
        utilidad_articulo=("utilidad", "sum"),
        cant_remitos=("remito_id", "nunique")
    ).reset_index()

    agrupado["utilidad_prom_unidad"] = (agrupado["utilidad_articulo"] / agrupado["cant_articulos"]).fillna(0.0).round(2)

    # Ordenar por Utilidad Descendente (Ranking)
    agrupado = agrupado.sort_values(by="utilidad_articulo", ascending=False).reset_index(drop=True)

    # Asignar Posición (1 a X) y dar formato
    agrupado["Posición"] = range(1, len(agrupado) + 1)
    agrupado["Artículo"] = agrupado["nro_articulo"]
    agrupado["Descripción"] = agrupado["descripcion"]
    agrupado["Rubro"] = agrupado["rubro_nombre"]
    agrupado["Cant. Remitos"] = agrupado["cant_remitos"].astype(int)
    agrupado["Cant. Artículos"] = agrupado["cant_articulos"].astype(int)
    agrupado["Utilidad ($)"] = agrupado["utilidad_articulo"].round(2)
    agrupado["Margen de Ganancia (%)"] = ((agrupado["utilidad_articulo"] / agrupado["venta_articulo"]) * 100.0).fillna(0.0).round(2)

    df_grilla = agrupado[["Posición", "Artículo", "Descripción", "Rubro", "Cant. Remitos", "Cant. Artículos", "Utilidad ($)", "Margen de Ganancia (%)"]]

    # Configuración de columnas en Streamlit dataframe
    column_cfg = {
        "Posición": st.column_config.NumberColumn("Posición", format="%d", width="small"),
        "Artículo": st.column_config.TextColumn("Artículo", width="small"),
        "Descripción": st.column_config.TextColumn("Descripción", width="medium"),
        "Rubro": st.column_config.TextColumn("Rubro", width="small"),
        "Cant. Remitos": st.column_config.NumberColumn("Cant. Remitos", width="small"),
        "Cant. Artículos": st.column_config.NumberColumn("Cant. Artículos", width="small"),
        "Utilidad ($)": st.column_config.NumberColumn("Utilidad ($)", width="medium"),
        "Margen de Ganancia (%)": st.column_config.NumberColumn("Margen de Ganancia (%)", width="medium"),
    }

    st.dataframe(
        df_grilla.style.format({
            "Posición": "{:d}",
            "Cant. Remitos": "{:d}",
            "Cant. Artículos": "{:d}",
            "Utilidad ($)": "$ {:,.2f}",
            "Margen de Ganancia (%)": "{:.2f} %"
        }),
        use_container_width=True,
        hide_index=True,
        height=390,
        column_config=column_cfg
    )

    # Totales del mes
    total_remitos_mes = int(agrupado["cant_remitos"].sum())
    total_articulos_mes = int(agrupado["cant_articulos"].sum())
    total_venta_mes = float(agrupado["venta_articulo"].sum())
    total_utilidad_mes = float(agrupado["utilidad_articulo"].sum())

    m_col0, m_col1, m_col2, m_col3 = st.columns([1.2, 1, 1, 1], gap="small")

    with m_col0:
        st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
        excel_bytes = generar_excel_articulos(agrupado, total_remitos_mes, total_articulos_mes, total_venta_mes, total_utilidad_mes)
        filename_excel = f"info_articulos_{anio_num:04d}{mes_num:02d}.xlsx"
        st.download_button(
            label="📊 Exportar a Excel",
            data=excel_bytes,
            file_name=filename_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_download_info_articulos"
        )

    with m_col1:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Remitos</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">{total_remitos_mes:,}</div>
        </div>
        """, unsafe_allow_html=True)

    with m_col2:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Artículos Vendidos</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">{total_articulos_mes:,}</div>
        </div>
        """, unsafe_allow_html=True)

    with m_col3:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Utilidad del Mes</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">$ {total_utilidad_mes:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # Explicación informativa de columnas al estilo Evidencia 2
    st.markdown("""
    <div style="font-size: 0.82rem; color: rgba(250, 250, 250, 0.75); background: rgba(255, 255, 255, 0.03); padding: 10px 14px; border-radius: 6px; border-left: 3px solid #29b6f6; margin-top: 15px; line-height: 1.5;">
        <div style="margin-bottom: 6px;"><b>Cant. Remitos</b>: Muestra la capilaridad de ventas. En cuántos remitos/pedidos distintos estuvo presente este producto (frecuencia de demanda en el mercado).</div>
        <div><b>Margen de Ganancia (%)</b>: Eficiencia del margen comercial del producto. Permite evaluar acuerdos de precios, descuentos y rentabilidad porcentual unitaria.</div>
    </div>
    """, unsafe_allow_html=True)

    # === SECCIÓN DE GRÁFICOS VISUALES ===
    st.markdown("---")
    st.subheader("📊 Análisis Gráfico del Mes")

    import altair as alt

    df_chart = agrupado.copy()
    df_chart["utilidad_fmt"] = df_chart["utilidad_articulo"].apply(lambda v: f"$ {v:,.2f}")
    df_chart["promedio_unidad_fmt"] = df_chart["utilidad_prom_unidad"].apply(lambda v: f"$ {v:,.2f}")
    df_chart["label_articulo"] = df_chart["nro_articulo"] + " - " + df_chart["descripcion"]

    tab_g1, tab_g2, tab_g3, tab_g4, tab_g5, tab_g6, tab_g7 = st.tabs([
        "💰 Utilidad por Artículo ($)", 
        "📦 Unidades Vendidas", 
        "🥧 Participación %",
        "🏷️ Utilidad por Rubro",
        "🏆 Top 10 vs Resto",
        "🏷️ Ganancia por Unidad ($/U)",
        "📊 Concentración Pareto %"
    ])

    def mostrar_explicacion_e(titulo, texto, color_borde="#ff4b4b"):
        st.markdown(f"""
        <div style="font-size: 0.82rem; color: rgba(250, 250, 250, 0.75); background: rgba(255, 255, 255, 0.03); padding: 8px 14px; border-radius: 6px; border-left: 3px solid {color_borde}; margin-top: 10px; line-height: 1.4;">
            <b>{titulo}</b>: {texto}
        </div>
        """, unsafe_allow_html=True)

    with tab_g1:
        chart_util = alt.Chart(df_chart).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#ff4b4b"
        ).encode(
            x=alt.X("utilidad_articulo:Q", title="Utilidad ($)"),
            y=alt.Y("label_articulo:N", title="Artículo", sort="-x", axis=alt.Axis(labelLimit=350)),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("nro_articulo:N", title="Artículo"),
                alt.Tooltip("descripcion:N", title="Descripción"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Unidades Vendidas")
            ]
        ).properties(
            height=max(300, len(df_chart) * 35)
        )
        st.altair_chart(chart_util, use_container_width=True)
        mostrar_explicacion_e("Utilidad por Artículo ($)", "Aporta valor al mostrar el ranking de productos ordenados por la ganancia neta total generada en el mes. Permite identificar de inmediato tus productos comercialmente más rentables.", "#ff4b4b")

    with tab_g2:
        chart_cant = alt.Chart(df_chart).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#00c49f"
        ).encode(
            x=alt.X("cant_articulos:Q", title="Unidades Vendidas", axis=alt.Axis(format="d")),
            y=alt.Y("label_articulo:N", title="Artículo", sort="-x", axis=alt.Axis(labelLimit=350)),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("nro_articulo:N", title="Artículo"),
                alt.Tooltip("descripcion:N", title="Descripción"),
                alt.Tooltip("cant_articulos:Q", title="Unidades Vendidas"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)")
            ]
        ).properties(
            height=max(300, len(df_chart) * 35)
        )
        st.altair_chart(chart_cant, use_container_width=True)
        mostrar_explicacion_e("Unidades Vendidas por Artículo", "Aporta valor al medir el volumen físico de salida del depósito. Permite detectar qué artículos tienen mayor rotación de stock independientemente de su precio.", "#00c49f")

    with tab_g3:
        df_chart_pie_raw = df_chart.sort_values(by="utilidad_articulo", ascending=False).copy()
        if len(df_chart_pie_raw) > 10:
            top10 = df_chart_pie_raw.iloc[:10].copy()
            resto_util = df_chart_pie_raw.iloc[10:]["utilidad_articulo"].sum()
            df_chart_pie = pd.concat([
                top10[["label_articulo", "utilidad_articulo"]],
                pd.DataFrame([{
                    "label_articulo": "Otros",
                    "utilidad_articulo": resto_util
                }])
            ], ignore_index=True)
        else:
            df_chart_pie = df_chart_pie_raw[["label_articulo", "utilidad_articulo"]].copy()

        df_chart_pie["utilidad_fmt"] = df_chart_pie["utilidad_articulo"].apply(lambda v: f"$ {v:,.2f}")
        df_chart_pie["porcentaje_utilidad"] = (df_chart_pie["utilidad_articulo"] / total_utilidad_mes * 100).round(1) if total_utilidad_mes > 0 else 0
        df_chart_pie["porcentaje_fmt"] = df_chart_pie["porcentaje_utilidad"].apply(lambda p: f"{p:.1f}%")

        tot_pie = df_chart_pie["utilidad_articulo"].sum()
        df_chart_pie["fraccion"] = df_chart_pie["utilidad_articulo"] / tot_pie if tot_pie > 0 else 0
        df_chart_pie["start_angle"] = 2 * np.pi * (df_chart_pie["fraccion"].cumsum() - df_chart_pie["fraccion"])
        df_chart_pie["end_angle"] = 2 * np.pi * df_chart_pie["fraccion"].cumsum()

        base_pie = alt.Chart(df_chart_pie).encode(
            theta=alt.Theta("start_angle:Q", scale=None),
            theta2=alt.Theta2("end_angle:Q"),
            color=alt.Color("label_articulo:N", title="Artículo", sort=None, scale=alt.Scale(scheme="category20")),
            tooltip=[
                alt.Tooltip("label_articulo:N", title="Artículo"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("porcentaje_fmt:N", title="Participación %")
            ]
        )

        main_arc = base_pie.transform_filter(
            alt.datum["label_articulo"] != "Otros"
        ).mark_arc(innerRadius=50, outerRadius=140, stroke="#1e1e1e", strokeWidth=2)

        otros_arc = base_pie.transform_filter(
            alt.datum["label_articulo"] == "Otros"
        ).mark_arc(innerRadius=65, outerRadius=165, stroke="#1e1e1e", strokeWidth=2)

        chart_pie = alt.layer(main_arc, otros_arc).properties(
            height=360
        )
        st.altair_chart(chart_pie, use_container_width=True)
        mostrar_explicacion_e("Participación % en Utilidad", "Aporta valor al visualizar la ponderación relativa del Top 10 de catálogo sobre las ganancias mensuales del negocio, destacando el bloque sobrante 'Otros'.", "#29b6f6")

    with tab_g4:
        df_rubro = df_chart.groupby("Rubro").agg(
            utilidad_rubro=("utilidad_articulo", "sum"),
            cant_rubro=("cant_articulos", "sum")
        ).reset_index()
        df_rubro["utilidad_fmt"] = df_rubro["utilidad_rubro"].apply(lambda v: f"$ {v:,.2f}")

        chart_rubro = alt.Chart(df_rubro).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#29b6f6"
        ).encode(
            x=alt.X("utilidad_rubro:Q", title="Utilidad ($)"),
            y=alt.Y("Rubro:N", title="Rubro / Categoría", sort="-x"),
            tooltip=[
                alt.Tooltip("Rubro:N", title="Rubro"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_rubro:Q", title="Unidades Vendidas")
            ]
        ).properties(
            height=max(250, len(df_rubro) * 40)
        )
        st.altair_chart(chart_rubro, use_container_width=True)
        mostrar_explicacion_e("Utilidad por Rubro ($)", "Aporta valor al agrupar la ganancia neta por categoría de producto (ej. Gorras, Sombreros), permitiendo identificar cuáles son las familias comercialmente más rentables para la empresa.", "#29b6f6")

    with tab_g5:
        if len(df_chart) > 10:
            top10 = df_chart.iloc[:10].copy()
            resto_util = df_chart.iloc[10:]["utilidad_articulo"].sum()
            resto_cant = df_chart.iloc[10:]["cant_articulos"].sum()
            
            df_top10 = pd.concat([
                top10[["label_articulo", "utilidad_articulo", "cant_articulos"]],
                pd.DataFrame([{
                    "label_articulo": "Resto del Catálogo",
                    "utilidad_articulo": resto_util,
                    "cant_articulos": resto_cant
                }])
            ], ignore_index=True)
        else:
            df_top10 = df_chart[["label_articulo", "utilidad_articulo", "cant_articulos"]].copy()

        df_top10["utilidad_fmt"] = df_top10["utilidad_articulo"].apply(lambda v: f"$ {v:,.2f}")

        chart_top10 = alt.Chart(df_top10).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5
        ).encode(
            x=alt.X("utilidad_articulo:Q", title="Utilidad ($)"),
            y=alt.Y("label_articulo:N", title="Artículo / Grupo", sort="-x", axis=alt.Axis(labelLimit=350)),
            color=alt.condition(
                alt.datum["label_articulo"] == "Resto del Catálogo",
                alt.value("#78909c"),
                alt.value("#ff9800")
            ),
            tooltip=[
                alt.Tooltip("label_articulo:N", title="Artículo / Grupo"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Unidades Vendidas")
            ]
        ).properties(
            height=360
        )
        st.altair_chart(chart_top10, use_container_width=True)
        mostrar_explicacion_e("Top 10 Artículos vs. Resto", "Aporta valor al contrastar el aporte económico del Top 10 de productos principales frente al resto de la oferta comercial de la empresa.", "#ff9800")

    with tab_g6:
        df_prom = df_chart.sort_values(by="utilidad_prom_unidad", ascending=False).copy()
        chart_prom_art = alt.Chart(df_prom).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#ab47bc"
        ).encode(
            x=alt.X("utilidad_prom_unidad:Q", title="Ganancia Promedio / Unidad ($)"),
            y=alt.Y("label_articulo:N", title="Artículo", sort="-x", axis=alt.Axis(labelLimit=350)),
            tooltip=[
                alt.Tooltip("nro_articulo:N", title="Artículo"),
                alt.Tooltip("descripcion:N", title="Descripción"),
                alt.Tooltip("promedio_unidad_fmt:N", title="Ganancia / Unidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Unidades Vendidas"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad Total ($)")
            ]
        ).properties(
            height=max(300, len(df_prom) * 35)
        )
        st.altair_chart(chart_prom_art, use_container_width=True)
        mostrar_explicacion_e("Ganancia Promedio por Unidad ($ / Unid)", "Aporta valor al revelar cuántos pesos netos de utilidad deja la venta de cada unidad individual, clave para analizar listas de precios y costos de reposición.", "#ab47bc")

    with tab_g7:
        df_pareto = df_chart.sort_values(by="utilidad_articulo", ascending=False).copy()
        df_pareto["utilidad_acum"] = df_pareto["utilidad_articulo"].cumsum()
        df_pareto["pareto_pct"] = (df_pareto["utilidad_acum"] / total_utilidad_mes * 100).round(1) if total_utilidad_mes > 0 else 0
        df_pareto["pareto_fmt"] = df_pareto["pareto_pct"].apply(lambda p: f"{p:.1f}%")

        chart_pareto = alt.Chart(df_pareto).mark_line(
            color="#7c4dff",
            point=alt.OverlayMarkDef(color="#7c4dff", size=60),
            strokeWidth=2
        ).encode(
            x=alt.X("nro_articulo:N", title="Artículo (por Ranking)", sort=None),
            y=alt.Y("pareto_pct:Q", title="Concentración Acumulada (%)", scale=alt.Scale(domain=[0, 100])),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("nro_articulo:N", title="Artículo"),
                alt.Tooltip("descripcion:N", title="Descripción"),
                alt.Tooltip("pareto_fmt:N", title="Acumulado %"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)")
            ]
        ).properties(
            height=340
        )

        # Línea horizontal de referencia en Y = 80% (roja punteada)
        rule_80_y = alt.Chart(pd.DataFrame({'y': [80.0]})).mark_rule(
            color='#ff4b4b',
            strokeDash=[4, 4],
            strokeWidth=2
        ).encode(y='y:Q')

        # Línea vertical de referencia en el artículo que alcanza el 80% (verde punteada)
        tot_articulos_catalogo = len(df_pareto)
        df_cutoff = df_pareto[df_pareto["pareto_pct"] >= 80.0]

        if not df_cutoff.empty:
            pos_corte = int(df_cutoff.iloc[0]["Posición"])
            articulo_corte_80 = df_cutoff.iloc[0]["nro_articulo"]
            pct_catalogo_corte = round((pos_corte / tot_articulos_catalogo) * 100, 1) if tot_articulos_catalogo > 0 else 0
            pct_util_corte = df_cutoff.iloc[0]["pareto_pct"]

            rule_80_x = alt.Chart(pd.DataFrame({'nro_articulo': [articulo_corte_80]})).mark_rule(
                color='#00e676',
                strokeDash=[4, 4],
                strokeWidth=2
            ).encode(x=alt.X('nro_articulo:N', sort=None))

            text_annotation = alt.Chart(pd.DataFrame([{
                'nro_articulo': articulo_corte_80,
                'y': 20.0,
                'texto': f"Corte 80% ({pct_catalogo_corte}% catálogo)"
            }])).mark_text(
                align='left',
                dx=6,
                dy=0,
                color='#00e676',
                fontSize=12,
                fontWeight='bold'
            ).encode(
                x=alt.X('nro_articulo:N', sort=None),
                y='y:Q',
                text='texto:N'
            )

            chart_pareto_final = chart_pareto + rule_80_y + rule_80_x + text_annotation
            explicacion_pareto = f"Aporta valor al aplicar el principio 80/20 de inventario ABC: El <b>{pct_catalogo_corte}%</b> de tu catálogo ({pos_corte} de {tot_articulos_catalogo} artículos) representa el <b>{pct_util_corte}%</b> de la utilidad total del mes."
        else:
            chart_pareto_final = chart_pareto + rule_80_y
            explicacion_pareto = "Aporta valor al aplicar el principio 80/20 de inventario ABC. Permite comprobar qué porcentaje exacto de tu catálogo representa el 80% de tus utilidades mensuales."

        st.altair_chart(chart_pareto_final, use_container_width=True)
        mostrar_explicacion_e("Análisis de Concentración Pareto (Acumulado %)", explicacion_pareto, "#7c4dff")

    st.markdown("---")
    st.markdown(f"`{config.FOOTER_APP}`")
