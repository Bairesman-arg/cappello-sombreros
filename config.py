# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl.utils import get_column_letter

VERSION = "1.1.56"
TITULO_APP = f"🧢 SISTEMA CAPELLO vs {VERSION}"
# TITULO_APP = "INTRODUCCION A PYTHON"
FOOTER_APP = "Sistema Capello® - Powered by Python and Streamlit - Telegram: @Bairesman - 2026"

RUBRO_DEFAULT = "GORRAS"
VENDEDOR_DEFAULT = "VICENTE"

# No estan en uso estás tres variables por el momento.
DEFAULT_FOLDER = "Escritorio/REMITOS CONSIGNACION"
DEFAULT_FOLDER_DATA = "DATA"
DEFAULT_FOLDER_REPORTS = "INFORMES"

# Funciones Comunes ---
def init_clientes_articulos():
    # Recarga Clientes y Articulos en diversas páginas
    try:
        if "porc_dto" in st.session_state:
            st.session_state.porc_dto = None
        del st.session_state.clientes_df
        del st.session_state.articulos_df
    except:
        pass

def get_desktop_path() -> str:
    """Obtiene la ruta del Escritorio del usuario."""
    return os.path.join(os.path.expanduser("~"), "Desktop")

def format_display_path(full_path: str) -> str:
    """
    Formatea la ruta para mostrar al usuario:
    - Convierte todas las barras '\\' a '/'
    - Si la ruta contiene 'Desktop' o 'Escritorio' (incluyendo carpetas redirigidas como OneDrive/Desktop),
      recorta el inicio y muestra a partir de 'Desktop/...' o 'Escritorio/...'
    """
    if not full_path:
        return ""
    clean_path = str(full_path).replace("\\", "/")
    parts = clean_path.split("/")
    for i, part in enumerate(parts):
        if part.lower() in ("desktop", "escritorio"):
            return "/".join(parts[i:])
    return clean_path

def df_to_excel_bytes(df: pd.DataFrame, drop_cols=None) -> bytes:
    """Convierte un DataFrame a bytes de Excel (.xlsx) filtrando columnas ID, formateando floats a 2 decimales (.00) y auto-ajustando anchos de columna."""
    df_export = df.copy()
    
    # 1. Eliminar columnas pasadas explícitamente (ej: 'Seleccionado')
    if drop_cols:
        for col in drop_cols:
            if col in df_export.columns:
                df_export = df_export.drop(columns=[col])
                
    # 2. Eliminar automáticamente cualquier columna 'id' o que termine en '_id'
    id_cols = [col for col in df_export.columns if col.lower() == 'id' or col.lower().endswith('_id')]
    if id_cols:
        df_export = df_export.drop(columns=id_cols)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Datos')
        ws = writer.sheets['Datos']
        
        # Formatear celdas numéricas flotantes con 2 decimales (.00)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float) and cell.value is not None:
                    cell.number_format = '0.00'
                elif isinstance(cell.value, (int,)) and cell.value is not None:
                    col_name = df_export.columns[cell.column - 1]
                    if pd.api.types.is_float_dtype(df_export[col_name]):
                        cell.number_format = '0.00'

        # Auto-ajustar ancho de cada columna al valor más largo + 4 caracteres
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value) if cell.value is not None else ""
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return output.getvalue()

def save_excel_with_folder_dialog(df: pd.DataFrame, filename_prefix: str = "Export", drop_cols=None):
    """Abre diálogo nativo de Windows (Tkinter) para seleccionar carpeta y guarda la planilla Excel (.xlsx), reemplazando si ya existe."""
    excel_bytes = df_to_excel_bytes(df, drop_cols=drop_cols)
    clean_prefix = filename_prefix if filename_prefix.startswith("#") else f"#{filename_prefix}"
    filename = f"{clean_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        
        initial_dir = os.path.join(get_desktop_path(), "EXCEL CAPELLO")
        if not os.path.exists(initial_dir):
            initial_dir = get_desktop_path()

        folder = filedialog.askdirectory(
            title=f"Seleccione la carpeta donde guardar {filename}",
            initialdir=initial_dir
        )
        root.destroy()
        
        if folder:
            dest_path = os.path.join(folder, filename)
            # Reemplazar si el archivo ya existe (modo wb de Python sobrescribe por defecto)
            with open(dest_path, "wb") as f_out:
                f_out.write(excel_bytes)
            
            display_path = format_display_path(dest_path)
            raw_prefix = filename_prefix.lstrip("#")
            st.session_state[f"excel_saved_msg_{filename_prefix}"] = f"🎉 Planilla de **{raw_prefix}** guardada con éxito en: **{display_path}**"
            st.rerun()
        else:
            st.warning("⚠️ No se seleccionó carpeta.")
    except Exception as e:
        st.error(f"❌ Error al guardar el archivo Excel: {e}")
