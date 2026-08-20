# -*- coding: utf-8 -*-
import datetime
import calendar
import pandas as pd
import streamlit as st
from sqlalchemy import text
from models import engine
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import numpy as np
import config

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def generar_excel_empresas(agrupado, total_remitos_mes, total_articulos_mes, total_venta_mes, total_utilidad_mes):
    output = io.BytesIO()
    df_excel = pd.DataFrame({
        "Posición": agrupado["Posición"],
        "Boca": agrupado["Boca"],
        "Nombre del Cliente": agrupado["Nombre del Cliente"],
        "Cant. Remitos": agrupado["Cant. Remitos"],
        "Cant. Artículos": agrupado["Cant. Artículos"],
        "Utilidad ($)": agrupado["utilidad_cliente"].round(2),
        "Margen de Ganancia (%)": agrupado["Margen de Ganancia (%)"]
    })
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Ranking por Empresa")
        ws = writer.sheets["Ranking por Empresa"]
        
        # Estilos openpyxl
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo pastel tenue
        bold_font = Font(bold=True)

        # Aplicar estilo al encabezado (Fila 1)
        for col_idx in range(1, 8):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = bold_font

        # Formatear datos
        for row in range(2, len(df_excel) + 2):
            ws.cell(row=row, column=1).number_format = '#,##0'
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=7).number_format = '0.00"%"'
        
        # Fila final de Totales en Negrita
        total_margen_mes = (total_utilidad_mes / total_venta_mes * 100.0) if total_venta_mes > 0 else 0.0
        tot_row = len(df_excel) + 3
        c1 = ws.cell(row=tot_row, column=1, value="Total del Mes")
        c4 = ws.cell(row=tot_row, column=4, value=total_remitos_mes)
        c4.number_format = '#,##0'
        c5 = ws.cell(row=tot_row, column=5, value=total_articulos_mes)
        c5.number_format = '#,##0'
        c6 = ws.cell(row=tot_row, column=6, value=total_utilidad_mes)
        c6.number_format = '"$"#,##0.00'
        c7 = ws.cell(row=tot_row, column=7, value=total_margen_mes)
        c7.number_format = '0.00"%"'

        top_border = Border(top=Side(style='thin'))
        for col_idx in range(1, 8):
            cell = ws.cell(row=tot_row, column=col_idx)
            cell.font = bold_font
            cell.border = top_border
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
    return output.getvalue()

def info_empresas_ranking():
    st.header("Informes - Ranking por Empresa")
    st.markdown('<div style="font-size: 0.85rem; color: #00E676; font-weight: 500; margin-top: -0.4rem; margin-bottom: 0.8rem;">El Análisis Gráfico se encuentra al final de la página</div>', unsafe_allow_html=True)
    st.markdown("---")

    # Fecha actual para predeterminar la selección
    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    # Controles de selección de Mes y Año
    col_mes, col_anio, _ = st.columns([2, 2, 4], gap="small")
    
    with col_mes:
        mes_nombre = st.selectbox(
            "Mes",
            options=MESES,
            index=current_month - 1,
            key="info_empresas_mes"
        )
        mes_num = MESES.index(mes_nombre) + 1

    with col_anio:
        min_year = min(2026, current_year)
        max_year = max(2050, current_year)
        anios_disponibles = list(range(min_year, max_year + 1))
        index_anio = anios_disponibles.index(current_year)
        anio_num = st.selectbox(
            "Año",
            options=anios_disponibles,
            index=index_anio,
            key="info_empresas_anio"
        )

    # Rango de fechas del mes completo seleccionado
    _, ult_dia = calendar.monthrange(anio_num, mes_num)
    fecha_inicio = datetime.date(anio_num, mes_num, 1)
    fecha_fin = datetime.date(anio_num, mes_num, ult_dia)

    # Consulta a base de datos de remitos con fecha_retiro en el mes seleccionado
    try:
        with engine.begin() as conn:
            query = text("""
                SELECT 
                    r.id AS remito_id,
                    r.fecha_retiro,
                    c.id AS cliente_id,
                    c.razon_social,
                    COALESCE(c.boca, 0) AS boca,
                    COALESCE(r.porc_dto, c.porc_dto, 0) AS porc_dto,
                    ri.entregados,
                    COALESCE(ri.devueltos, 0) AS devueltos,
                    COALESCE(ri.precio_real_item, a.precio_real, 0) AS precio_real,
                    COALESCE(a.costo, 0) AS costo
                FROM remitos r
                JOIN clientes c ON r.cliente_id = c.id
                JOIN remito_items ri ON r.id = ri.remito_id
                JOIN articulos a ON ri.articulo_id = a.id
                WHERE r.fecha_retiro IS NOT NULL 
                  AND r.fecha_retiro >= :start_date 
                  AND r.fecha_retiro <= :end_date
            """)
            df_raw = pd.read_sql(query, conn, params={"start_date": fecha_inicio, "end_date": fecha_fin})
    except Exception as e:
        st.error(f"Error al consultar la base de datos: {e}")
        return

    if df_raw.empty:
        st.info(f"ℹ️ No existen Remitos con ventas registradas en {mes_nombre} {anio_num}.")
        st.metric("Total Utilidad del Mes", "$ 0.00")
        return

    # Aplicar cálculo de vendidos y utilidad ítem por ítem
    def calcular_item(row):
        entregados = int(row["entregados"]) if pd.notna(row["entregados"]) else 0
        devueltos = int(row["devueltos"]) if pd.notna(row["devueltos"]) else 0
        vendidos = max(0, entregados - devueltos)

        p_real = float(row["precio_real"]) if pd.notna(row["precio_real"]) else 0.0
        porc_dto = float(row["porc_dto"]) if pd.notna(row["porc_dto"]) else 0.0
        costo = float(row["costo"]) if pd.notna(row["costo"]) else 0.0

        if p_real > 0 and vendidos > 0:
            p_dto = p_real * (1.0 - (porc_dto / 100.0))
            venta_total = p_dto * vendidos
            utilidad = (p_dto - costo) * vendidos
        else:
            venta_total = 0.0
            utilidad = 0.0

        return pd.Series({"vendidos": vendidos, "venta_total": venta_total, "utilidad": utilidad})

    res_items = df_raw.apply(calcular_item, axis=1)
    df_raw["vendidos"] = res_items["vendidos"]
    df_raw["venta_total"] = res_items["venta_total"]
    df_raw["utilidad"] = res_items["utilidad"]

    # Agrupar por Cliente (razon_social, boca)
    agrupado = df_raw.groupby(["cliente_id", "razon_social", "boca"]).agg(
        cant_articulos=("vendidos", "sum"),
        venta_cliente=("venta_total", "sum"),
        utilidad_cliente=("utilidad", "sum"),
        cant_remitos=("remito_id", "nunique")
    ).reset_index()

    agrupado["utilidad_prom_remito"] = (agrupado["utilidad_cliente"] / agrupado["cant_remitos"]).round(2)

    # Ordenar por Utilidad Descendente (Ranking)
    agrupado = agrupado.sort_values(by="utilidad_cliente", ascending=False).reset_index(drop=True)

    # Asignar Posición (1 a X)
    agrupado["Posición"] = range(1, len(agrupado) + 1)
    agrupado["Boca"] = agrupado["boca"].astype(int)
    agrupado["Nombre del Cliente"] = agrupado["razon_social"]
    agrupado["Cant. Remitos"] = agrupado["cant_remitos"].astype(int)
    agrupado["Cant. Artículos"] = agrupado["cant_articulos"].astype(int)
    agrupado["Utilidad ($)"] = agrupado["utilidad_cliente"].round(2)
    agrupado["Margen de Ganancia (%)"] = ((agrupado["utilidad_cliente"] / agrupado["venta_cliente"]) * 100.0).fillna(0.0).round(2)

    df_grilla = agrupado[["Posición", "Boca", "Nombre del Cliente", "Cant. Remitos", "Cant. Artículos", "Utilidad ($)", "Margen de Ganancia (%)"]]

    # Configuración de columnas en Streamlit dataframe
    column_cfg = {
        "Posición": st.column_config.NumberColumn("Posición", format="%d", width="small"),
        "Boca": st.column_config.NumberColumn("Boca", format="%d", width="small"),
        "Nombre del Cliente": st.column_config.TextColumn("Nombre del Cliente", width="medium"),
        "Cant. Remitos": st.column_config.NumberColumn("Cant. Remitos", width="small"),
        "Cant. Artículos": st.column_config.NumberColumn("Cant. Artículos", width="small"),
        "Utilidad ($)": st.column_config.NumberColumn("Utilidad ($)", width="medium"),
        "Margen de Ganancia (%)": st.column_config.NumberColumn("Margen de Ganancia (%)", width="medium"),
    }

    st.dataframe(
        df_grilla.style.format({
            "Posición": "{:d}",
            "Boca": "{:d}",
            "Cant. Remitos": "{:d}",
            "Cant. Artículos": "{:d}",
            "Utilidad ($)": "$ {:,.2f}",
            "Margen de Ganancia (%)": "{:.2f} %"
        }),
        use_container_width=True,
        hide_index=True,
        height=390,
        column_config=column_cfg
    )

    # Totales del mes
    total_remitos_mes = int(agrupado["cant_remitos"].sum())
    total_articulos_mes = int(agrupado["cant_articulos"].sum())
    total_venta_mes = float(agrupado["venta_cliente"].sum())
    total_utilidad_mes = float(agrupado["utilidad_cliente"].sum())

    m_col0, m_col1, m_col2, m_col3 = st.columns([1.2, 1, 1, 1], gap="small")

    with m_col0:
        st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
        excel_bytes = generar_excel_empresas(agrupado, total_remitos_mes, total_articulos_mes, total_venta_mes, total_utilidad_mes)
        filename_excel = f"info_empresas_{anio_num:04d}{mes_num:02d}.xlsx"
        st.download_button(
            label="📊 Exportar a Excel",
            data=excel_bytes,
            file_name=filename_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_download_info_empresas"
        )

    with m_col1:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Remitos</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">{total_remitos_mes:,}</div>
        </div>
        """, unsafe_allow_html=True)

    with m_col2:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Artículos Vendidos</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">{total_articulos_mes:,}</div>
        </div>
        """, unsafe_allow_html=True)

    with m_col3:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Utilidad del Mes</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">$ {total_utilidad_mes:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # Explicación informativa de columnas al estilo Evidencia 2
    st.markdown("""
    <div style="font-size: 0.82rem; color: rgba(250, 250, 250, 0.75); background: rgba(255, 255, 255, 0.03); padding: 10px 14px; border-radius: 6px; border-left: 3px solid #29b6f6; margin-top: 15px; line-height: 1.5;">
        <div style="margin-bottom: 6px;"><b>Cant. Remitos</b>: Frecuencia del cliente. Muestra cuántos pedidos/remitos realizó el cliente en el mes, permitiendo diferenciar compras recurrentes de pedidos únicos.</div>
        <div><b>Margen de Ganancia (%)</b>: Rentabilidad porcentual que deja cada cliente, clave para evaluar descuentos concedidos y acuerdos comerciales.</div>
    </div>
    """, unsafe_allow_html=True)

    # === SECCIÓN DE GRÁFICOS VISUALES ===
    st.markdown("---")
    st.subheader("📊 Análisis Gráfico del Mes")

    import altair as alt

    df_chart = agrupado.copy()
    df_chart["utilidad_fmt"] = df_chart["utilidad_cliente"].apply(lambda v: f"$ {v:,.2f}")
    df_chart["promedio_remito_fmt"] = df_chart["utilidad_prom_remito"].apply(lambda v: f"$ {v:,.2f}")

    tab_g1, tab_g2, tab_g3, tab_g4, tab_g5, tab_g6, tab_g7 = st.tabs([
        "💰 Utilidad por Empresa ($)", 
        "📦 Artículos por Empresa", 
        "🥧 Participación %",
        "🎯 Cantidad vs Utilidad",
        "🏆 Top 10 vs Resto",
        "🧾 Promedio por Remito ($)",
        "📊 Acumulado Pareto %"
    ])

    def mostrar_explicacion_e(titulo, texto, color_borde="#ff4b4b"):
        st.markdown(f"""
        <div style="font-size: 0.82rem; color: rgba(250, 250, 250, 0.75); background: rgba(255, 255, 255, 0.03); padding: 8px 14px; border-radius: 6px; border-left: 3px solid {color_borde}; margin-top: 10px; line-height: 1.4;">
            <b>{titulo}</b>: {texto}
        </div>
        """, unsafe_allow_html=True)

    with tab_g1:
        chart_util = alt.Chart(df_chart).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#ff4b4b"
        ).encode(
            x=alt.X("utilidad_cliente:Q", title="Utilidad ($)"),
            y=alt.Y("Nombre del Cliente:N", title="Cliente", sort="-x", axis=alt.Axis(labelLimit=350)),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("Nombre del Cliente:N", title="Cliente"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos")
            ]
        ).properties(
            height=max(300, len(df_chart) * 35)
        )
        st.altair_chart(chart_util, use_container_width=True)
        mostrar_explicacion_e("Utilidad por Empresa ($)", "Muestra el ranking de clientes ordenados por la ganancia total que aportaron en el mes. Permite identificar a tus clientes comercialmente más rentables.", "#ff4b4b")

    with tab_g2:
        chart_cant = alt.Chart(df_chart).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#00c49f"
        ).encode(
            x=alt.X("cant_articulos:Q", title="Artículos Vendidos", axis=alt.Axis(format="d")),
            y=alt.Y("Nombre del Cliente:N", title="Cliente", sort="-x", axis=alt.Axis(labelLimit=350)),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("Nombre del Cliente:N", title="Cliente"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)")
            ]
        ).properties(
            height=max(300, len(df_chart) * 35)
        )
        st.altair_chart(chart_cant, use_container_width=True)
        mostrar_explicacion_e("Artículos por Empresa", "Muestra la cantidad total de unidades compradas por cada cliente en el período. Permite evaluar cuáles son tus clientes de mayor volumen físico.", "#00c49f")

    with tab_g3:
        df_chart_pie_raw = df_chart.sort_values(by="utilidad_cliente", ascending=False).copy()
        if len(df_chart_pie_raw) > 10:
            top10 = df_chart_pie_raw.iloc[:10].copy()
            resto_util = df_chart_pie_raw.iloc[10:]["utilidad_cliente"].sum()
            df_chart_pie = pd.concat([
                top10[["Nombre del Cliente", "utilidad_cliente"]],
                pd.DataFrame([{
                    "Nombre del Cliente": "Otros",
                    "utilidad_cliente": resto_util
                }])
            ], ignore_index=True)
        else:
            df_chart_pie = df_chart_pie_raw[["Nombre del Cliente", "utilidad_cliente"]].copy()

        df_chart_pie["utilidad_fmt"] = df_chart_pie["utilidad_cliente"].apply(lambda v: f"$ {v:,.2f}")
        df_chart_pie["porcentaje_utilidad"] = (df_chart_pie["utilidad_cliente"] / total_utilidad_mes * 100).round(1) if total_utilidad_mes > 0 else 0
        df_chart_pie["porcentaje_fmt"] = df_chart_pie["porcentaje_utilidad"].apply(lambda p: f"{p:.1f}%")

        # Calcular posiciones angulares unificadas para mantener la torta 100% coordinada
        tot_pie = df_chart_pie["utilidad_cliente"].sum()
        df_chart_pie["fraccion"] = df_chart_pie["utilidad_cliente"] / tot_pie if tot_pie > 0 else 0
        df_chart_pie["start_angle"] = 2 * np.pi * (df_chart_pie["fraccion"].cumsum() - df_chart_pie["fraccion"])
        df_chart_pie["end_angle"] = 2 * np.pi * df_chart_pie["fraccion"].cumsum()

        base_pie = alt.Chart(df_chart_pie).encode(
            theta=alt.Theta("start_angle:Q", scale=None),
            theta2=alt.Theta2("end_angle:Q"),
            color=alt.Color("Nombre del Cliente:N", title="Cliente", sort=None, scale=alt.Scale(scheme="category20")),
            tooltip=[
                alt.Tooltip("Nombre del Cliente:N", title="Cliente"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("porcentaje_fmt:N", title="Participación %")
            ]
        )

        main_arc = base_pie.transform_filter(
            alt.datum["Nombre del Cliente"] != "Otros"
        ).mark_arc(innerRadius=50, outerRadius=140, stroke="#1e1e1e", strokeWidth=2)

        otros_arc = base_pie.transform_filter(
            alt.datum["Nombre del Cliente"] == "Otros"
        ).mark_arc(innerRadius=65, outerRadius=165, stroke="#1e1e1e", strokeWidth=2)

        chart_pie = alt.layer(main_arc, otros_arc).properties(
            height=360
        )
        st.altair_chart(chart_pie, use_container_width=True)
        mostrar_explicacion_e("Participación % en Utilidad", "Representa la porción que aporta cada empresa sobre la ganancia total del mes. Muestra individualmente a los 10 primeros clientes y destaca la porción 'Otros' desprendida de la torta.", "#29b6f6")

    with tab_g4:
        chart_scatter = alt.Chart(df_chart).mark_circle(size=140, opacity=0.85, color="#1e88e5").encode(
            x=alt.X("cant_articulos:Q", title="Cantidad de Artículos", axis=alt.Axis(format="d")),
            y=alt.Y("utilidad_cliente:Q", title="Utilidad ($)"),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("Nombre del Cliente:N", title="Cliente"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)")
            ]
        ).properties(
            height=340
        )

        chart_scatter_text = chart_scatter.mark_text(
            align='left',
            baseline='middle',
            dx=10,
            fontSize=11,
            color='#ffffff'
        ).encode(
            text='Nombre del Cliente:N'
        )

        st.altair_chart(chart_scatter + chart_scatter_text, use_container_width=True)
        mostrar_explicacion_e("Matriz de Rentabilidad (Cantidad vs. Utilidad $)", "Relaciona el volumen de artículos frente a la ganancia obtenida. Permite detectar clientes VIP/Alto Margen (baja cantidad, alta ganancia) vs. clientes de Alto Volumen (muchas unidades pero menor margen).", "#1e88e5")

    with tab_g5:
        if len(df_chart) > 10:
            top10 = df_chart.iloc[:10].copy()
            resto_util = df_chart.iloc[10:]["utilidad_cliente"].sum()
            resto_cant = df_chart.iloc[10:]["cant_articulos"].sum()
            
            df_top10 = pd.concat([
                top10[["Nombre del Cliente", "utilidad_cliente", "cant_articulos"]],
                pd.DataFrame([{
                    "Nombre del Cliente": "Resto del Mercado",
                    "utilidad_cliente": resto_util,
                    "cant_articulos": resto_cant
                }])
            ], ignore_index=True)
        else:
            df_top10 = df_chart[["Nombre del Cliente", "utilidad_cliente", "cant_articulos"]].copy()

        df_top10["utilidad_fmt"] = df_top10["utilidad_cliente"].apply(lambda v: f"$ {v:,.2f}")

        chart_top10 = alt.Chart(df_top10).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5
        ).encode(
            x=alt.X("utilidad_cliente:Q", title="Utilidad ($)"),
            y=alt.Y("Nombre del Cliente:N", title="Cliente / Grupo", sort="-x", axis=alt.Axis(labelLimit=350)),
            color=alt.condition(
                alt.datum["Nombre del Cliente"] == "Resto del Mercado",
                alt.value("#78909c"),
                alt.value("#ff9800")
            ),
            tooltip=[
                alt.Tooltip("Nombre del Cliente:N", title="Cliente / Grupo"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos")
            ]
        ).properties(
            height=360
        )
        st.altair_chart(chart_top10, use_container_width=True)
        mostrar_explicacion_e("Top 10 Clientes vs. Resto", "Compara en una sola vista la ganancia generada por tus 10 clientes principales frente a la suma del resto del mercado.", "#ff9800")

    with tab_g6:
        df_prom = df_chart.sort_values(by="utilidad_prom_remito", ascending=False).copy()
        chart_prom_emp = alt.Chart(df_prom).mark_bar(
            cornerRadiusTopRight=5,
            cornerRadiusBottomRight=5,
            color="#ab47bc"
        ).encode(
            x=alt.X("utilidad_prom_remito:Q", title="Utilidad Promedio / Remito ($)"),
            y=alt.Y("Nombre del Cliente:N", title="Cliente", sort="-x", axis=alt.Axis(labelLimit=350)),
            tooltip=[
                alt.Tooltip("Nombre del Cliente:N", title="Cliente"),
                alt.Tooltip("promedio_remito_fmt:N", title="Promedio / Remito ($)"),
                alt.Tooltip("cant_remitos:Q", title="Remitos Operados"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad Total ($)")
            ]
        ).properties(
            height=max(300, len(df_prom) * 35)
        )
        st.altair_chart(chart_prom_emp, use_container_width=True)
        mostrar_explicacion_e("Utilidad Promedio por Remito ($ / Remito)", "Muestra qué clientes compran mediante operaciones/remitos grandes (alto valor unitario) frente a clientes que hacen muchos pedidos pequeños.", "#ab47bc")

    with tab_g7:
        df_pareto = df_chart.sort_values(by="utilidad_cliente", ascending=False).copy()
        df_pareto["utilidad_acum"] = df_pareto["utilidad_cliente"].cumsum()
        df_pareto["pareto_pct"] = (df_pareto["utilidad_acum"] / total_utilidad_mes * 100).round(1) if total_utilidad_mes > 0 else 0
        df_pareto["pareto_fmt"] = df_pareto["pareto_pct"].apply(lambda p: f"{p:.1f}%")

        chart_pareto = alt.Chart(df_pareto).mark_line(
            color="#7c4dff",
            point=alt.OverlayMarkDef(color="#7c4dff", size=60),
            strokeWidth=2
        ).encode(
            x=alt.X("Nombre del Cliente:N", title="Cliente (por Ranking)", sort=None),
            y=alt.Y("pareto_pct:Q", title="Concentración Acumulada (%)", scale=alt.Scale(domain=[0, 100])),
            tooltip=[
                alt.Tooltip("Posición:Q", title="Posición"),
                alt.Tooltip("Nombre del Cliente:N", title="Cliente"),
                alt.Tooltip("pareto_fmt:N", title="Acumulado %"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)")
            ]
        ).properties(
            height=340
        )

        # Línea horizontal de referencia en Y = 80% (roja punteada)
        rule_80_y = alt.Chart(pd.DataFrame({'y': [80.0]})).mark_rule(
            color='#ff4b4b',
            strokeDash=[4, 4],
            strokeWidth=2
        ).encode(y='y:Q')

        # Línea vertical de referencia en el cliente que alcanza el 80% (verde punteada)
        tot_clientes = len(df_pareto)
        df_cutoff = df_pareto[df_pareto["pareto_pct"] >= 80.0]

        if not df_cutoff.empty:
            pos_corte = int(df_cutoff.iloc[0]["Posición"])
            cliente_corte_80 = df_cutoff.iloc[0]["Nombre del Cliente"]
            pct_clientes_corte = round((pos_corte / tot_clientes) * 100, 1) if tot_clientes > 0 else 0
            pct_util_corte = df_cutoff.iloc[0]["pareto_pct"]

            rule_80_x = alt.Chart(pd.DataFrame({'Nombre del Cliente': [cliente_corte_80]})).mark_rule(
                color='#00e676',
                strokeDash=[4, 4],
                strokeWidth=2
            ).encode(x=alt.X('Nombre del Cliente:N', sort=None))

            text_annotation = alt.Chart(pd.DataFrame([{
                'Nombre del Cliente': cliente_corte_80,
                'y': 20.0,
                'texto': f"Corte 80% ({pct_clientes_corte}% empresas)"
            }])).mark_text(
                align='left',
                dx=6,
                dy=0,
                color='#00e676',
                fontSize=12,
                fontWeight='bold'
            ).encode(
                x=alt.X('Nombre del Cliente:N', sort=None),
                y='y:Q',
                text='texto:N'
            )

            chart_pareto_final = chart_pareto + rule_80_y + rule_80_x + text_annotation
            explicacion_pareto = f"Muestra la curva acumulada de participación comercial: El <b>{pct_clientes_corte}%</b> de tus clientes ({pos_corte} de {tot_clientes} empresas) genera el <b>{pct_util_corte}%</b> de las ganancias del mes."
        else:
            chart_pareto_final = chart_pareto + rule_80_y
            explicacion_pareto = "Muestra la curva acumulada de participación comercial (de 0% a 100%). Permite verificar si se cumple la Regla 80/20 (el 20% de las empresas genera el 80% de tus ingresos)."

        st.altair_chart(chart_pareto_final, use_container_width=True)
        mostrar_explicacion_e("Análisis de Concentración Pareto (Acumulado %)", explicacion_pareto, "#7c4dff")

    st.markdown("---")
    st.markdown(f"`{config.FOOTER_APP}`")
