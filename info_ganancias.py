# -*- coding: utf-8 -*-
import datetime
import calendar
import pandas as pd
import streamlit as st
from sqlalchemy import text
from models import engine
import io
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import config

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

def generar_excel_ganancias(agrupado, total_remitos_mes, total_articulos_mes, total_venta_mes, total_utilidad_mes):
    output = io.BytesIO()
    df_excel = pd.DataFrame({
        "Fecha": agrupado["Fecha"],
        "Cant. Remitos": agrupado["Cant. Remitos"],
        "Artículos Vendidos": agrupado["Artículos Vendidos"],
        "Venta Total ($)": agrupado["venta_dia"].round(2),
        "Utilidad ($)": agrupado["utilidad_dia"].round(2),
        "Margen de Ganancia (%)": agrupado["Margen de Ganancia (%)"]
    })
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_excel.to_excel(writer, index=False, sheet_name="Ganancias por Día")
        ws = writer.sheets["Ganancias por Día"]
        
        # Estilos openpyxl
        header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo pastel tenue
        bold_font = Font(bold=True)

        # Aplicar estilo al encabezado (Fila 1)
        for col_idx in range(1, 7):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = bold_font

        # Formatear datos
        for row in range(2, len(df_excel) + 2):
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
            ws.cell(row=row, column=6).number_format = '0.00"%"'
        
        # Fila final de Totales en Negrita
        total_margen_mes = (total_utilidad_mes / total_venta_mes * 100.0) if total_venta_mes > 0 else 0.0
        tot_row = len(df_excel) + 3
        c1 = ws.cell(row=tot_row, column=1, value="Total del Mes")
        c2 = ws.cell(row=tot_row, column=2, value=total_remitos_mes)
        c2.number_format = '#,##0'
        c3 = ws.cell(row=tot_row, column=3, value=total_articulos_mes)
        c3.number_format = '#,##0'
        c4 = ws.cell(row=tot_row, column=4, value=total_venta_mes)
        c4.number_format = '"$"#,##0.00'
        c5 = ws.cell(row=tot_row, column=5, value=total_utilidad_mes)
        c5.number_format = '"$"#,##0.00'
        c6 = ws.cell(row=tot_row, column=6, value=total_margen_mes)
        c6.number_format = '0.00"%"'

        top_border = Border(top=Side(style='thin'))
        for c in (c1, c2, c3, c4, c5, c6):
            c.font = bold_font
            c.border = top_border
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)
            
    return output.getvalue()

def info_ganancias_dia():
    st.header("Informes - Ganancias por Día")
    st.markdown('<div style="font-size: 0.85rem; color: #00E676; font-weight: 500; margin-top: -0.4rem; margin-bottom: 0.8rem;">El Análisis Gráfico se encuentra al final de la página</div>', unsafe_allow_html=True)
    st.markdown("---")

    # Fecha actual para predeterminar la selección
    today = datetime.date.today()
    current_year = today.year
    current_month = today.month

    # Controles de selección de Mes y Año
    col_mes, col_anio, _ = st.columns([2, 2, 4], gap="small")
    
    with col_mes:
        mes_nombre = st.selectbox(
            "Mes",
            options=MESES,
            index=current_month - 1,
            key="info_ganancias_mes"
        )
        mes_num = MESES.index(mes_nombre) + 1

    with col_anio:
        min_year = min(2026, current_year)
        max_year = max(2050, current_year)
        anios_disponibles = list(range(min_year, max_year + 1))
        index_anio = anios_disponibles.index(current_year)
        anio_num = st.selectbox(
            "Año",
            options=anios_disponibles,
            index=index_anio,
            key="info_ganancias_anio"
        )

    # Rango de fechas del mes completo seleccionado
    _, ult_dia = calendar.monthrange(anio_num, mes_num)
    fecha_inicio = datetime.date(anio_num, mes_num, 1)
    fecha_fin = datetime.date(anio_num, mes_num, ult_dia)

    # Consulta a base de datos de remitos con fecha_retiro en el mes seleccionado
    try:
        with engine.begin() as conn:
            query = text("""
                SELECT 
                    r.id AS remito_id,
                    r.fecha_retiro,
                    COALESCE(r.porc_dto, c.porc_dto, 0) AS porc_dto,
                    ri.entregados,
                    COALESCE(ri.devueltos, 0) AS devueltos,
                    COALESCE(ri.precio_real_item, a.precio_real, 0) AS precio_real,
                    COALESCE(a.costo, 0) AS costo
                FROM remitos r
                JOIN clientes c ON r.cliente_id = c.id
                JOIN remito_items ri ON r.id = ri.remito_id
                JOIN articulos a ON ri.articulo_id = a.id
                WHERE r.fecha_retiro IS NOT NULL 
                  AND r.fecha_retiro >= :start_date 
                  AND r.fecha_retiro <= :end_date
                ORDER BY r.fecha_retiro ASC
            """)
            df_raw = pd.read_sql(query, conn, params={"start_date": fecha_inicio, "end_date": fecha_fin})
    except Exception as e:
        st.error(f"Error al consultar la base de datos: {e}")
        return

    if df_raw.empty:
        st.info(f"ℹ️ No existen Remitos con ventas registradas en {mes_nombre} {anio_num}.")
        st.metric("Total Utilidad del Mes", "$ 0.00")
        return

    # Asegurar conversión de fecha a date
    df_raw["fecha_retiro_dt"] = pd.to_datetime(df_raw["fecha_retiro"]).dt.date

    # Aplicar cálculo de vendidos y utilidad ítem por ítem
    def calcular_item(row):
        entregados = int(row["entregados"]) if pd.notna(row["entregados"]) else 0
        devueltos = int(row["devueltos"]) if pd.notna(row["devueltos"]) else 0
        vendidos = max(0, entregados - devueltos)

        p_real = float(row["precio_real"]) if pd.notna(row["precio_real"]) else 0.0
        porc_dto = float(row["porc_dto"]) if pd.notna(row["porc_dto"]) else 0.0
        costo = float(row["costo"]) if pd.notna(row["costo"]) else 0.0

        if p_real > 0 and vendidos > 0:
            p_dto = p_real * (1.0 - (porc_dto / 100.0))
            venta_total = p_dto * vendidos
            utilidad = (p_dto - costo) * vendidos
        else:
            venta_total = 0.0
            utilidad = 0.0

        return pd.Series({"vendidos": vendidos, "venta_total": venta_total, "utilidad": utilidad})

    res_items = df_raw.apply(calcular_item, axis=1)
    df_raw["vendidos"] = res_items["vendidos"]
    df_raw["venta_total"] = res_items["venta_total"]
    df_raw["utilidad"] = res_items["utilidad"]

    # Agrupar por fecha_retiro cronológicamente
    agrupado = df_raw.groupby("fecha_retiro_dt").agg(
        cant_articulos=("vendidos", "sum"),
        venta_dia=("venta_total", "sum"),
        utilidad_dia=("utilidad", "sum"),
        cant_remitos=("remito_id", "nunique")
    ).reset_index()

    agrupado = agrupado.sort_values(by="fecha_retiro_dt", ascending=True)

    # Formatear columnas para la grilla
    agrupado["Fecha"] = agrupado["fecha_retiro_dt"].apply(lambda d: d.strftime("%d/%m/%Y"))
    agrupado["Cant. Remitos"] = agrupado["cant_remitos"].astype(int)
    agrupado["Artículos Vendidos"] = agrupado["cant_articulos"].astype(int)
    agrupado["Venta Total ($)"] = agrupado["venta_dia"].round(2)
    agrupado["Utilidad ($)"] = agrupado["utilidad_dia"].round(2)
    agrupado["Margen de Ganancia (%)"] = ((agrupado["utilidad_dia"] / agrupado["venta_dia"]) * 100.0).fillna(0.0).round(2)
    agrupado["utilidad_prom_remito"] = (agrupado["utilidad_dia"] / agrupado["cant_remitos"]).round(2)

    df_grilla = agrupado[["Fecha", "Cant. Remitos", "Artículos Vendidos", "Venta Total ($)", "Utilidad ($)", "Margen de Ganancia (%)"]]

    # Configuración de columnas en Streamlit dataframe
    column_cfg = {
        "Fecha": st.column_config.TextColumn("Fecha", width="small"),
        "Cant. Remitos": st.column_config.NumberColumn("Cant. Remitos", width="small"),
        "Artículos Vendidos": st.column_config.NumberColumn("Cant. Artículos", width="small"),
        "Venta Total ($)": st.column_config.NumberColumn("Venta Total ($)", width="medium"),
        "Utilidad ($)": st.column_config.NumberColumn("Utilidad ($)", width="medium"),
        "Margen de Ganancia (%)": st.column_config.NumberColumn("Margen de Ganancia (%)", width="medium"),
    }

    st.dataframe(
        df_grilla.style.format({
            "Cant. Remitos": "{:d}",
            "Artículos Vendidos": "{:d}",
            "Venta Total ($)": "$ {:,.2f}",
            "Utilidad ($)": "$ {:,.2f}",
            "Margen de Ganancia (%)": "{:.2f} %"
        }),
        use_container_width=True,
        hide_index=True,
        height=390,
        column_config=column_cfg
    )

    # Totales del mes
    total_remitos_mes = int(agrupado["Cant. Remitos"].sum())
    total_articulos_mes = int(agrupado["Artículos Vendidos"].sum())
    total_venta_mes = float(agrupado["venta_dia"].sum())
    total_utilidad_mes = float(agrupado["utilidad_dia"].sum())

    m_col0, m_col1, m_col2, m_col3 = st.columns([1.2, 1, 1, 1], gap="small")

    with m_col0:
        st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
        excel_bytes = generar_excel_ganancias(agrupado, total_remitos_mes, total_articulos_mes, total_venta_mes, total_utilidad_mes)
        filename_excel = f"info_ganancias_{anio_num:04d}{mes_num:02d}.xlsx"
        st.download_button(
            label="📊 Exportar a Excel",
            data=excel_bytes,
            file_name=filename_excel,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="btn_download_info_ganancias"
        )

    with m_col1:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Remitos</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">{total_remitos_mes:,}</div>
        </div>
        """, unsafe_allow_html=True)

    with m_col2:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Venta del Mes</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">$ {total_venta_mes:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    with m_col3:
        st.markdown(f"""
        <div style="text-align: right; width: 100%;">
            <div style="font-size: 0.85rem; color: rgba(250, 250, 250, 0.7); font-weight: 400; margin-bottom: 4px;">Total Utilidad del Mes</div>
            <div style="font-size: 1.8rem; font-weight: 600; color: var(--text-color, #ffffff); line-height: 1.2;">$ {total_utilidad_mes:,.2f}</div>
        </div>
        """, unsafe_allow_html=True)

    # Explicación informativa de columnas al estilo Evidencia 2
    st.markdown("""
    <div style="font-size: 0.82rem; color: rgba(250, 250, 250, 0.75); background: rgba(255, 255, 255, 0.03); padding: 10px 14px; border-radius: 6px; border-left: 3px solid #29b6f6; margin-top: 15px; line-height: 1.5;">
        <div style="margin-bottom: 6px;"><b>Cant. Remitos</b>: Densidad del día. No es lo mismo vender 47 arts. en 2 remitos grandes (pocos clientes, mayor volumen por cliente) que en 20 remitos chicos (alta carga operativa de despacho).</div>
        <div style="margin-bottom: 6px;"><b>Venta Total ($)</b>: Ver la Venta Total junto a la Utilidad permite analizar de un vistazo cuánto ingresó vs. cuánto quedó de ganancia neta.</div>
        <div><b>Margen de Ganancia (%)</b>: Muestra la eficiencia del margen comercial del día (por ejemplo, si ese día se vendieron productos con mayor o menor margen).</div>
    </div>
    """, unsafe_allow_html=True)

    # === SECCIÓN DE GRÁFICOS VISUALES ===
    st.markdown("---")
    st.subheader("📊 Análisis Gráfico del Mes")

    import altair as alt

    df_chart = agrupado.copy()
    df_chart["utilidad_fmt"] = df_chart["utilidad_dia"].apply(lambda v: f"$ {v:,.2f}")
    df_chart["utilidad_acum"] = df_chart["utilidad_dia"].cumsum()
    df_chart["utilidad_acum_fmt"] = df_chart["utilidad_acum"].apply(lambda v: f"$ {v:,.2f}")
    df_chart["promedio_remito_fmt"] = df_chart["utilidad_prom_remito"].apply(lambda v: f"$ {v:,.2f}")

    tab_g1, tab_g2, tab_g3, tab_g4, tab_g5, tab_g6 = st.tabs([
        "💰 Utilidad Diaria ($)", 
        "📦 Artículos Vendidos", 
        "📈 Evolución Acumulada ($)",
        "📅 Día de la Semana",
        "📆 Semana del Mes",
        "🧾 Promedio por Remito ($)"
    ])

    def mostrar_explicacion_g(titulo, texto, color_borde="#ff4b4b"):
        st.markdown(f"""
        <div style="font-size: 0.82rem; color: rgba(250, 250, 250, 0.75); background: rgba(255, 255, 255, 0.03); padding: 8px 14px; border-radius: 6px; border-left: 3px solid {color_borde}; margin-top: 10px; line-height: 1.4;">
            <b>{titulo}</b>: {texto}
        </div>
        """, unsafe_allow_html=True)

    with tab_g1:
        chart_util = alt.Chart(df_chart).mark_bar(
            cornerRadiusTopLeft=5,
            cornerRadiusTopRight=5,
            color="#ff4b4b"
        ).encode(
            x=alt.X("Fecha:N", title="Fecha (DD/MM/AAAA)", sort=None),
            y=alt.Y("utilidad_dia:Q", title="Utilidad ($)"),
            tooltip=[
                alt.Tooltip("Fecha:N", title="Fecha"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos")
            ]
        ).properties(
            height=340
        )
        st.altair_chart(chart_util, use_container_width=True)
        mostrar_explicacion_g("Utilidad Diaria ($)", "Muestra la ganancia neta obtenida en cada día del mes por las entregas y retiros registrados. Permite detectar picos y caídas en la rentabilidad diaria.", "#ff4b4b")

    with tab_g2:
        chart_cant = alt.Chart(df_chart).mark_bar(
            cornerRadiusTopLeft=5,
            cornerRadiusTopRight=5,
            color="#00c49f"
        ).encode(
            x=alt.X("Fecha:N", title="Fecha (DD/MM/AAAA)", sort=None),
            y=alt.Y("cant_articulos:Q", title="Artículos Vendidos", axis=alt.Axis(format="d")),
            tooltip=[
                alt.Tooltip("Fecha:N", title="Fecha"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)")
            ]
        ).properties(
            height=340
        )
        st.altair_chart(chart_cant, use_container_width=True)
        mostrar_explicacion_g("Artículos Vendidos", "Muestra el volumen total de unidades comercializadas cada día. Permite correlacionar los días de mayor movimiento físico con la utilidad generada.", "#00c49f")

    with tab_g3:
        chart_acum_line = alt.Chart(df_chart).mark_area(
            color=alt.Gradient(
                gradient='linear',
                stops=[alt.GradientStop(color='#1e88e5', offset=1), alt.GradientStop(color='rgba(30, 136, 229, 0.1)', offset=0)],
                x1=1, x2=1, y1=1, y2=0
            ),
            line={'color': '#29b6f6', 'strokeWidth': 2}
        ).encode(
            x=alt.X("Fecha:N", title="Fecha (DD/MM/AAAA)", sort=None),
            y=alt.Y("utilidad_acum:Q", title="Utilidad Acumulada ($)"),
            tooltip=[
                alt.Tooltip("Fecha:N", title="Fecha"),
                alt.Tooltip("utilidad_acum_fmt:N", title="Utilidad Acumulada ($)"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad del Día")
            ]
        ).properties(
            height=340
        )
        st.altair_chart(chart_acum_line, use_container_width=True)
        mostrar_explicacion_g("Evolución Acumulada ($)", "Muestra la curva de crecimiento progresivo de las ganancias a lo largo del mes. Permite proyectar el cumplimiento de metas mensuales.", "#29b6f6")

    with tab_g4:
        dias_es = {0: "1-Lunes", 1: "2-Martes", 2: "3-Miércoles", 3: "4-Jueves", 4: "5-Viernes", 5: "6-Sábado", 6: "7-Domingo"}
        df_chart["dia_semana_num"] = df_chart["fecha_retiro_dt"].apply(lambda d: d.weekday())
        df_chart["Día de Semana"] = df_chart["dia_semana_num"].map(dias_es)

        df_dia_semana = df_chart.groupby("Día de Semana").agg(
            utilidad_total=("utilidad_dia", "sum"),
            cant_articulos=("cant_articulos", "sum")
        ).reset_index().sort_values(by="Día de Semana")

        df_dia_semana["Día"] = df_dia_semana["Día de Semana"].apply(lambda s: s.split("-")[1])
        df_dia_semana["utilidad_fmt"] = df_dia_semana["utilidad_total"].apply(lambda v: f"$ {v:,.2f}")

        chart_dow = alt.Chart(df_dia_semana).mark_bar(
            cornerRadiusTopLeft=5,
            cornerRadiusTopRight=5,
            color="#ab47bc"
        ).encode(
            x=alt.X("Día:N", title="Día de la Semana", sort=["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]),
            y=alt.Y("utilidad_total:Q", title="Utilidad ($)"),
            tooltip=[
                alt.Tooltip("Día:N", title="Día"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos")
            ]
        ).properties(
            height=340
        )
        st.altair_chart(chart_dow, use_container_width=True)
        mostrar_explicacion_g("Rendimiento por Día de la Semana", "Agrupa los ingresos según el día de la semana (Lunes a Domingo). Permite identificar los días de mayor actividad comercial para optimizar logística y cobranza.", "#ab47bc")

    with tab_g5:
        def obtener_semana(dia_num):
            if dia_num <= 7:
                return "S1 (Días 1-7)"
            elif dia_num <= 14:
                return "S2 (Días 8-14)"
            elif dia_num <= 21:
                return "S3 (Días 15-21)"
            else:
                return "S4 (Días 22+)"

        df_chart["Semana"] = df_chart["fecha_retiro_dt"].apply(lambda d: obtener_semana(d.day))

        df_semana = df_chart.groupby("Semana").agg(
            utilidad_total=("utilidad_dia", "sum"),
            cant_articulos=("cant_articulos", "sum")
        ).reset_index()

        df_semana["utilidad_fmt"] = df_semana["utilidad_total"].apply(lambda v: f"$ {v:,.2f}")

        chart_sem = alt.Chart(df_semana).mark_bar(
            cornerRadiusTopLeft=5,
            cornerRadiusTopRight=5,
            color="#26a69a"
        ).encode(
            x=alt.X("Semana:N", title="Semana del Mes", sort=None),
            y=alt.Y("utilidad_total:Q", title="Utilidad ($)"),
            tooltip=[
                alt.Tooltip("Semana:N", title="Semana"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad ($)"),
                alt.Tooltip("cant_articulos:Q", title="Artículos Vendidos")
            ]
        ).properties(
            height=340
        )
        st.altair_chart(chart_sem, use_container_width=True)
        mostrar_explicacion_g("Rendimiento por Semana del Mes", "Compara el desempeño financiero por tramos semanales del mes. Permite evaluar el comportamiento de compra entre la 1ª y la 2ª quincena.", "#26a69a")

    with tab_g6:
        chart_prom = alt.Chart(df_chart).mark_line(
            color="#ff9800",
            point=alt.OverlayMarkDef(color="#ff9800", size=60),
            strokeWidth=2
        ).encode(
            x=alt.X("Fecha:N", title="Fecha (DD/MM/AAAA)", sort=None),
            y=alt.Y("utilidad_prom_remito:Q", title="Utilidad Promedio / Remito ($)"),
            tooltip=[
                alt.Tooltip("Fecha:N", title="Fecha"),
                alt.Tooltip("promedio_remito_fmt:N", title="Promedio / Remito ($)"),
                alt.Tooltip("cant_remitos:Q", title="Cantidad Remitos del Día"),
                alt.Tooltip("utilidad_fmt:N", title="Utilidad Total del Día")
            ]
        ).properties(
            height=340
        )
        st.altair_chart(chart_prom, use_container_width=True)
        mostrar_explicacion_g("Utilidad Promedio por Remito ($)", "Muestra el valor medio de ganancia por cada operación o remito procesado en el día. Permite distinguir entre días de operaciones de alto valor frente a días de pedidos menores.", "#ff9800")

    st.markdown("---")
    st.markdown(f"`{config.FOOTER_APP}`")
