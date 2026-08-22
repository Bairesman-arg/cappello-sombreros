# -*- coding: utf-8 -*-
import io
import datetime
from datetime import datetime as dt_class
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from sqlalchemy import text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import config
from models import engine, get_remito_completo

def format_fecha(val):
    """Formatea la fecha a formato DD/MM/YYYY o retorna string vacío si no existe o es nula."""
    if pd.isna(val) or val is None or str(val).strip() in ["", "NaT", "None", "nan"]:
        return ""
    try:
        dt = pd.to_datetime(val)
        if pd.isna(dt):
            return ""
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return ""

@st.cache_data(show_spinner=False)
def get_consultas_df():
    """Consulta la base de datos para obtener el listado completo de remitos."""
    query = text("""
        SELECT 
            r.id AS "Nro. Remito",
            c.boca AS "Nro. Boca",
            c.razon_social AS "Razón Social",
            COALESCE(SUM(ri.entregados), 0) AS "Cant. Artículos",
            r.fecha_entrega AS "Fecha Entrega",
            r.fecha_retiro AS "Fecha Retiro",
            r.porc_dto AS "% Dto",
            r.observaciones AS "Observaciones"
        FROM remitos r
        JOIN clientes c ON r.cliente_id = c.id
        LEFT JOIN remito_items ri ON r.id = ri.remito_id
        GROUP BY r.id, c.boca, c.razon_social, r.fecha_entrega, r.fecha_retiro, r.porc_dto, r.observaciones
        ORDER BY r.id DESC;
    """)
    with engine.connect() as conn:
        df = pd.read_sql(query, conn)
    
    if not df.empty:
        # Asegurar ordenamiento descendente por Nro. Remito
        df = df.sort_values(by="Nro. Remito", ascending=False).reset_index(drop=True)
        
        # Casteo y limpieza de datos
        df["Nro. Remito"] = df["Nro. Remito"].fillna(0).astype(int)
        df["Nro. Boca"] = df["Nro. Boca"].fillna(0).astype(int)
        df["Razón Social"] = df["Razón Social"].fillna("").astype(str)
        df["Cant. Artículos"] = df["Cant. Artículos"].fillna(0).astype(int)
        df["% Dto"] = df["% Dto"].fillna(0.0).astype(float)
        df["Observaciones"] = df["Observaciones"].fillna("").astype(str).replace(["None", "none", "nan", "NaN"], "")
        
        # Formatear fechas como texto DD/MM/YYYY para evitar que aparezca 'None' cuando sean nulas
        df["Fecha Entrega"] = df["Fecha Entrega"].apply(format_fecha)
        df["Fecha Retiro"] = df["Fecha Retiro"].apply(format_fecha)

    return df

@st.cache_data(show_spinner=False)
def get_items_remito_df_and_utility(remito_id):
    """Obtiene el detalle de artículos de un remito y calcula la utilidad total del remito."""
    datos = get_remito_completo(remito_id)
    if not datos or "items" not in datos or datos["items"].empty:
        return pd.DataFrame(), 0.0
    
    cabecera = datos.get("cabecera", {})
    porc_dto_val = float(cabecera.get("porc_dto", 0) or 0)
    
    df_items = datos["items"].copy()
    
    resumen_df = pd.DataFrame()
    resumen_df["Nro. Artículo"] = df_items["nro_articulo"].astype(str)
    resumen_df["Descripción"] = df_items["descripcion"].astype(str)
    resumen_df["Precio Real"] = df_items["precio_real"].fillna(0.0).astype(float)
    resumen_df["Entregados"] = df_items["entregados"].fillna(0).astype(int)
    resumen_df["Devueltos"] = df_items["devueltos"].fillna(0).astype(int)
    resumen_df["Vendidos"] = (resumen_df["Entregados"] - resumen_df["Devueltos"]).apply(lambda x: max(0, x))
    resumen_df["Observaciones"] = df_items["observaciones"].fillna("").astype(str).replace(["None", "none", "nan", "NaN"], "")
    
    # Calcular Utilidad del Remito
    utilidades = []
    for idx in df_items.index:
        pr = float(df_items.loc[idx, "precio_real"] or 0)
        co = float(df_items.loc[idx, "costo"] or 0)
        ve = max(0, int(df_items.loc[idx, "entregados"] or 0) - int(df_items.loc[idx, "devueltos"] or 0))
        if pr > 0 and ve > 0:
            p_dto = pr * (1.0 - (porc_dto_val / 100.0))
            utilidades.append((p_dto - co) * ve)
        else:
            utilidades.append(0.0)
    
    utilidad_total = float(sum(utilidades))
    return resumen_df, utilidad_total

@st.cache_data(show_spinner=False)
def generar_excel_remitos(df_remitos):
    """Genera el libro Excel TodosLosRemitos_aaammdd.xlsx con cabeceras de remitos e ítems anidados en columna 2 (en 1 consulta optimizada)."""
    if df_remitos.empty:
        return b""

    remito_ids = df_remitos["Nro. Remito"].tolist()
    
    # Consultar todos los ítems de los remitos seleccionados en UNA consulta masiva optimizada
    if remito_ids:
        placeholders = ",".join([str(rid) for rid in remito_ids])
        query_items = text(f"""
            SELECT 
                ri.remito_id,
                a.nro_articulo AS "Nro. Artículo",
                a.descripcion AS "Descripción",
                COALESCE(ri.precio_real_item, a.precio_real, 0) AS "Precio Real",
                COALESCE(ri.entregados, 0) AS "Entregados",
                COALESCE(ri.devueltos, 0) AS "Devueltos",
                COALESCE(ri.observaciones_item, '') AS "Observaciones"
            FROM remito_items ri
            JOIN articulos a ON ri.articulo_id = a.id
            WHERE ri.remito_id IN ({placeholders})
            ORDER BY ri.remito_id DESC, ri.id ASC;
        """)
        with engine.connect() as conn:
            all_items_df = pd.read_sql(query_items, conn)
    else:
        all_items_df = pd.DataFrame()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todos los Remitos"

    yellow_header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    yellow_item_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11, bold=False)

    main_headers = ["Nro. Remito", "Nro. Boca", "Razón Social", "Cant. Artículos", "Fecha Entrega", "Fecha Retiro", "% Dto", "Observaciones"]
    item_headers = ["Nro. Artículo", "Descripción", "Precio Real", "Entregados", "Devueltos", "Vendidos", "Observaciones"]

    row_idx = 1
    for _, remito_row in df_remitos.iterrows():
        remito_id = int(remito_row["Nro. Remito"])
        
        # 1. Cabecera Principal del Remito (repetida para cada remito)
        for col_num, h_text in enumerate(main_headers, start=1):
            cell = ws.cell(row=row_idx, column=col_num, value=h_text)
            cell.fill = yellow_header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center" if col_num in [1, 2, 4, 7] else "left", vertical="center")
        row_idx += 1

        # 2. Fila de datos del Remito (Columnas 1 a 8 / A a H)
        ws.cell(row=row_idx, column=1, value=remito_id).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=int(remito_row["Nro. Boca"])).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=str(remito_row["Razón Social"]))
        ws.cell(row=row_idx, column=4, value=int(remito_row["Cant. Artículos"])).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=5, value=str(remito_row["Fecha Entrega"])).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=str(remito_row["Fecha Retiro"])).alignment = Alignment(horizontal="center")
        
        try:
            dto_val = float(remito_row["% Dto"])
            dto_str = f"{dto_val:.0f}%"
        except (ValueError, TypeError):
            dto_str = "0%"
        ws.cell(row=row_idx, column=7, value=dto_str).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=8, value=str(remito_row["Observaciones"]))
        
        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).font = regular_font

        row_idx += 1

        # Ítems de este remito
        if not all_items_df.empty:
            sub_items = all_items_df[all_items_df["remito_id"] == remito_id]
            if not sub_items.empty:
                # Cabecera de ítems arrancando en Columna 2 (Columna B)
                for c_offset, header_text in enumerate(item_headers, start=2):
                    cell = ws.cell(row=row_idx, column=c_offset, value=header_text)
                    cell.fill = yellow_item_fill
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal="center" if c_offset in [4, 5, 6, 7] else "left", vertical="center")
                row_idx += 1

                # Filas de ítems en Columnas B a H
                for _, item_row in sub_items.iterrows():
                    entregados = int(item_row["Entregados"])
                    devueltos = int(item_row["Devueltos"])
                    vendidos = max(0, entregados - devueltos)

                    ws.cell(row=row_idx, column=2, value=str(item_row["Nro. Artículo"])).font = regular_font
                    ws.cell(row=row_idx, column=3, value=str(item_row["Descripción"])).font = regular_font
                    
                    ic4 = ws.cell(row=row_idx, column=4, value=float(item_row["Precio Real"]))
                    ic4.number_format = '"$"#,##0.00'
                    ic4.font = regular_font
                    
                    ws.cell(row=row_idx, column=5, value=entregados).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=6, value=devueltos).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=7, value=vendidos).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=8, value=str(item_row["Observaciones"])).font = regular_font
                    
                    for c in range(2, 9):
                        ws.cell(row=row_idx, column=c).font = regular_font

                    row_idx += 1
        
        # Fila vacía entre remitos
        row_idx += 1

    col_widths = {'A': 14, 'B': 16, 'C': 35, 'D': 16, 'E': 15, 'F': 15, 'G': 10, 'H': 40}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def remitos_consultas():
    st.title(config.TITULO_APP)
    st.header("Consulta de Remitos")

    st.markdown("`Seleccione la primera columna de la grilla de remitos para ver sus artículos.`")

    df = get_consultas_df()

    if df.empty:
        st.info("No hay remitos registrados en el sistema.")
    else:
        search_ver = st.session_state.get("consultas_search_version", 0)
        current_query = st.session_state.get("consultas_search_query", "")

        col_search, col_clear, col_filtrar, col_excel = st.columns([2.85, 0.35, 1, 1.1], gap="small")

        with col_search:
            search_input = st.text_input(
                "Filtrar",
                value=current_query,
                placeholder="Ingrese una Boca, Una Razón Social o parte, o una fecha de Entrega o Retiro...",
                label_visibility="collapsed",
                key=f"input_search_remitos_{search_ver}"
            )

        with col_clear:
            is_clear_disabled = not (str(current_query).strip() or str(search_input).strip())
            if st.button("↩️", key="btn_clear_search_remitos", help="Limpiar búsqueda", width="stretch", disabled=is_clear_disabled):
                st.session_state["consultas_search_query"] = ""
                st.session_state["consultas_search_version"] = search_ver + 1
                st.session_state["consultas_selected_idx"] = 0
                st.session_state["consultas_loading_filter"] = True
                st.rerun()

        with col_filtrar:
            if st.button("Filtrar", width="stretch", type="primary", key="btn_filtrar_remitos"):
                st.session_state["consultas_search_query"] = search_input
                st.session_state["consultas_loading_filter"] = True
                st.rerun()

        filter_loading_placeholder = st.empty()
        if st.session_state.get("consultas_loading_filter", False):
            filter_loading_placeholder.markdown(
                """
                <div style='background-color: #052c16; border: 1px solid #065f46; border-radius: 6px; padding: 10px 14px; color: #34d399; font-weight: 500; font-family: "Source Code Pro", Consolas, monospace; font-size: 0.95rem; margin-top: 10px; margin-bottom: 10px; width: 100%;'>
                    Un momento por favor...
                </div>
                """,
                unsafe_allow_html=True
            )

        # Filtrado inteligente sobre el DataFrame (Boca, Razón Social, Fecha Entrega, Fecha Retiro)
        query_text = search_input.strip().lower()
        if query_text:
            df_filtered = df[
                df["Nro. Boca"].astype(str).str.lower().str.contains(query_text, na=False) |
                df["Razón Social"].astype(str).str.lower().str.contains(query_text, na=False) |
                df["Fecha Entrega"].astype(str).str.lower().str.contains(query_text, na=False) |
                df["Fecha Retiro"].astype(str).str.lower().str.contains(query_text, na=False)
            ].reset_index(drop=True)
            estado_grilla = "filtrados"
        else:
            df_filtered = df
            estado_grilla = "totales"

        # Generar Excel optimizado para el listado actual con nombre #TodosLosRemitos_aaaammdd.xlsx
        excel_filename = f"#TodosLosRemitos_{dt_class.now().strftime('%Y%m%d')}.xlsx"
        excel_bytes = generar_excel_remitos(df_filtered if not df_filtered.empty else df)

        with col_excel:
            st.download_button(
                label="📊 Enviar a Excel",
                data=excel_bytes,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                key="btn_excel_remitos"
            )

        filter_loading_placeholder.empty()
        st.session_state["consultas_loading_filter"] = False

        st.subheader(f"Remitos ({len(df_filtered)} {estado_grilla})")

        if df_filtered.empty:
            st.warning("No se encontraron remitos que coincidan con la búsqueda.")
        else:
            # Grilla principal de remitos
            curr_selected_idx = st.session_state.get("consultas_selected_idx", 0)
            if curr_selected_idx >= len(df_filtered):
                curr_selected_idx = 0
                st.session_state["consultas_selected_idx"] = 0

            def highlight_selected_row(row):
                if row.name == curr_selected_idx:
                    return ['background-color: rgba(255, 75, 75, 0.28); font-weight: bold; color: #ffffff;'] * len(row)
                return [''] * len(row)

            styled_df = df_filtered.style.apply(highlight_selected_row, axis=1)

            column_config = {
                "Nro. Remito": st.column_config.NumberColumn("Nro. Remito", format="%d", width=105),
                "Nro. Boca": st.column_config.NumberColumn("Nro. Boca", format="%d", width=90),
                "Razón Social": st.column_config.TextColumn("Razón Social", width=280),
                "Cant. Artículos": st.column_config.NumberColumn("Cant. Artículos", format="%d", width=110),
                "Fecha Entrega": st.column_config.TextColumn("Fecha Entrega", width=120),
                "Fecha Retiro": st.column_config.TextColumn("Fecha Retiro", width=115),
                "% Dto": st.column_config.NumberColumn("% Dto", format="%.0f%%", width=70),
                "Observaciones": st.column_config.TextColumn("Observaciones", width=360),
            }

            num_rows = len(df_filtered)
            rows_to_show = min(max(num_rows, 1), 10)
            grid_height = int(38 + (rows_to_show * 35) + 1)

            selection_event = st.dataframe(
                styled_df,
                column_config=column_config,
                hide_index=True,
                width="content",
                height=grid_height,
                on_select="rerun",
                selection_mode="single-row",
                key="grid_consultas_remitos_main"
            )

            selected_rows = selection_event.selection.get("rows", [])
            if selected_rows and len(selected_rows) > 0 and selected_rows[0] < len(df_filtered):
                new_idx = selected_rows[0]
                if new_idx != curr_selected_idx:
                    st.session_state["consultas_selected_idx"] = new_idx
                    st.session_state["consultas_loading"] = True
                    st.rerun()

            selected_idx = st.session_state.get("consultas_selected_idx", 0)
            if selected_idx >= len(df_filtered):
                selected_idx = 0

            selected_remito_id = int(df_filtered.iloc[selected_idx]["Nro. Remito"])
            selected_cliente = df_filtered.iloc[selected_idx]["Razón Social"]

            loading_placeholder = st.empty()
            if st.session_state.get("consultas_loading", False):
                loading_placeholder.markdown(
                    """
                    <div style='background-color: #052c16; border: 1px solid #065f46; border-radius: 6px; padding: 10px 14px; color: #34d399; font-weight: 500; font-family: "Source Code Pro", Consolas, monospace; font-size: 0.95rem; margin-top: 15px; margin-bottom: 15px; width: 100%;'>
                        Recuperando artículos. Un momento por favor...
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            items_df, utilidad_remito = get_items_remito_df_and_utility(selected_remito_id)

            loading_placeholder.empty()
            st.session_state["consultas_loading"] = False

            st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
            st.subheader(f"Artículos del Remito #{selected_remito_id} - {selected_cliente}")

            if items_df.empty:
                st.info(f"El Remito #{selected_remito_id} no contiene artículos cargados.")
            else:
                items_column_config = {
                    "Nro. Artículo": st.column_config.TextColumn("Nro. Artículo", width=110),
                    "Descripción": st.column_config.TextColumn("Descripción", width=250),
                    "Precio Real": st.column_config.NumberColumn("Precio Real", format="$ %.2f", width=110),
                    "Entregados": st.column_config.NumberColumn("Entregados", format="%d", width=95),
                    "Devueltos": st.column_config.NumberColumn("Devueltos", format="%d", width=95),
                    "Vendidos": st.column_config.NumberColumn("Vendidos", format="%d", width=95),
                    "Observaciones": st.column_config.TextColumn("Observaciones", width=495),
                }

                num_items = len(items_df)
                items_rows_to_show = min(max(num_items, 1), 10)
                items_grid_height = int(38 + (items_rows_to_show * 35) + 1)

                st.dataframe(
                    items_df,
                    column_config=items_column_config,
                    hide_index=True,
                    width="content",
                    height=items_grid_height,
                    key=f"grid_items_remito_{selected_remito_id}"
                )

                t_entregados = int(items_df["Entregados"].sum())
                t_devueltos = int(items_df["Devueltos"].sum())
                t_vendidos = int(items_df["Vendidos"].sum())

                m1, m2, m3, m4 = st.columns(4)
                with m1:
                    st.metric("Total Entregados", f"{t_entregados}")
                with m2:
                    st.metric("Total Devueltos", f"{t_devueltos}")
                with m3:
                    st.metric("Total Vendidos", f"{t_vendidos}")
                with m4:
                    st.metric("Utilidad del Remito", f"$ {utilidad_remito:,.2f}")

    # Footer
    st.markdown(f"`{config.FOOTER_APP}`")

if __name__ == "__main__":
    remitos_consultas()
