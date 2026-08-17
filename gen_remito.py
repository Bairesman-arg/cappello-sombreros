import os
import io
import streamlit as st
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from models import get_remito_completo
import config

def gen_remito(remito_id: int, is_retiro=False) -> io.BytesIO:
    """
    Genera un archivo Excel del remito usando la plantilla REMITO_Master.xls
    y devuelve un buffer listo para descargar.
    """
    # Obtener datos desde BD
    data = get_remito_completo(remito_id)
    if not data:
        raise ValueError("Remito no encontrado")

    cab = data["cabecera"]
    items = data["items"]

    # Cargar plantilla
    template_path = os.path.join(os.path.dirname(__file__), "DOCS", "REMITO_Master.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    # --- Cabecera ---
    ws["A3"] = st.secrets["DIRECCION_CLIENTE"]
    ws["A5"] = cab["razon_social"]
    ws["H5"] = cab["boca"] or ""
    ws["H8"] = "Nro." + f'{remito_id:05d}'
    ws["A6"] = f"{cab['direccion'] or ''} - {cab['localidad'] or ''}"
    ws["G6"] = cab["telefono"] or ""
    if not is_retiro:  # Solo cuando se genera el Remito (1ra. vez)
        ws["H2"] = (1.0 - (float(cab["porc_dto"]) / 100.0)) if cab.get("porc_dto") else 1.0
    else:
        ws["H2"] = (1.0 - (float(cab["porc_dto"]) / 100.0)) if cab.get("porc_dto") else 1.0

    # --- Items ---
    base_row = 10
    for i, row in items.iterrows():
        ws[f"A{base_row+i}"] = row["nro_articulo"]
        ws[f"B{base_row+i}"] = row["descripcion"]
        # Columna C utiliza la fórmula nativa de la plantilla Excel: =IF(D10 ="","",D10*$H$2)
        ws[f"D{base_row+i}"] = float(row["precio_real"])
        ws[f"E{base_row+i}"] = int(row["entregados"])
        if is_retiro:
            ws[f"F{base_row+i}"] = int(row["devueltos"])
            ws[f"G{base_row+i}"] = int(row["entregados"]-row["devueltos"])
            ws[f"H{base_row+i}"] = row["observaciones"]

    # --- Fecha de entrega ---
    fecha = pd.to_datetime(cab["fecha_entrega"])
    ws["E45"] = fecha.day
    ws["F45"] = fecha.month
    ws["G45"] = fecha.year % 100

    # --- Fecha Retiro ---
    if is_retiro:
        try:
            fecha = pd.to_datetime(cab["fecha_retiro"]) #--
            ws["E46"] = fecha.day
            ws["F46"] = fecha.month
            ws["G46"] = fecha.year % 100
        except:
            pass

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def get_desktop_path() -> str:
    """Obtiene la ruta real del Escritorio de Windows, soportando redirecciones (como OneDrive)."""
    try:
        import ctypes
        buf = ctypes.create_unicode_buffer(260)
        # CSIDL_DESKTOP = 0
        if ctypes.windll.shell32.SHGetFolderPathW(None, 0, None, 0, buf) == 0 and buf.value:
            return buf.value
    except Exception:
        pass

    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders')
        val, _ = winreg.QueryValueEx(key, 'Desktop')
        expanded = os.path.expandvars(val)
        if expanded and os.path.exists(expanded):
            return expanded
    except Exception:
        pass

    user_profile = os.environ.get("USERPROFILE", os.path.expanduser("~"))
    try:
        for subfolder in os.listdir(user_profile):
            if "OneDrive" in subfolder:
                onedrive_desktop = os.path.join(user_profile, subfolder, "Desktop")
                if os.path.exists(onedrive_desktop):
                    return onedrive_desktop
    except Exception:
        pass

    return os.path.join(user_profile, "Desktop")

def is_local_app() -> bool:
    """Detecta si la aplicación se está ejecutando en un entorno Windows local con interfaz gráfica."""
    if os.name != 'nt':
        return False
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        root.destroy()
        return True
    except Exception:
        return False

def select_folder_native(default_dir=None) -> str:
    """Abre un cuadro de diálogo emergente de Windows para seleccionar carpeta."""
    try:
        import tkinter as tk
        from tkinter import filedialog

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        
        initial_dir = default_dir or os.path.join(get_desktop_path(), "REMITOS CONSIGNACION")
        if not os.path.exists(initial_dir):
            initial_dir = get_desktop_path()

        folder = filedialog.askdirectory(
            title="Seleccione la carpeta donde guardar el Remito",
            initialdir=initial_dir
        )
        root.destroy()
        return folder if folder else None
    except Exception:
        return None

def get_remito_filename(remito_id: int, is_retiro: bool = False) -> str:
    """
    Construye el nombre del archivo Excel en formato:
    Remito_<boca:04d>_<remito_id>.xlsx
    o bien
    Remito_<boca:04d>_<remito_id>_Ventas.xlsx (si es retiro)
    """
    boca_str = "0000"
    try:
        data = get_remito_completo(remito_id)
        if data and "cabecera" in data:
            boca_val = data["cabecera"].get("boca")
            if boca_val is not None and str(boca_val).strip() != "":
                boca_str = f"{int(boca_val):04d}"
    except Exception:
        pass

    if is_retiro:
        return f"Remito_{boca_str}_{remito_id}_Ventas.xlsx"
    else:
        return f"Remito_{boca_str}_{remito_id}.xlsx"

def save_remito_to_custom_folder(remito_id: int, folder_path: str, is_retiro: bool = False) -> str:
    """Guarda el remito Excel en la carpeta indicada por el usuario."""
    os.makedirs(folder_path, exist_ok=True)
    file_name = get_remito_filename(remito_id, is_retiro=is_retiro)
    target_path = os.path.join(folder_path, file_name)
    excel_buffer = gen_remito(remito_id, is_retiro=is_retiro)
    with open(target_path, "wb") as f:
        f.write(excel_buffer.getvalue())

    return target_path

def save_remito_to_desktop(remito_id: int, is_retiro: bool = False) -> str:
    """
    Genera el remito Excel y lo guarda directamente en Escritorio/REMITOS CONSIGNACION.
    Crea la carpeta si no existe.
    Devuelve la ruta absoluta del archivo guardado.
    """
    desktop_path = get_desktop_path()
    target_dir = os.path.join(desktop_path, "REMITOS CONSIGNACION")
    return save_remito_to_custom_folder(remito_id, target_dir, is_retiro=is_retiro)

def process_generate_remito(remito_id: int, is_retiro: bool = False, default_dir: str = None):
    """
    Función híbrida para procesar la generación del remito.
    Si se ejecuta localmente, abre el diálogo para elegir carpeta y guarda el archivo.
    Devuelve (guardado_exitoso: bool, mensaje_o_ruta: str, carpeta_seleccionada: str)
    """
    if is_local_app():
        selected_folder = select_folder_native(default_dir=default_dir)
        if selected_folder:
            saved_path = save_remito_to_custom_folder(remito_id, selected_folder, is_retiro=is_retiro)
            display_path = config.format_display_path(saved_path)
            return True, display_path, selected_folder
        else:
            return False, "Operación cancelada por el usuario.", None
    else:
        target_path = save_remito_to_desktop(remito_id, is_retiro=is_retiro)
        display_path = config.format_display_path(target_path)
        return True, display_path, None
